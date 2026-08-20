# -*- coding: utf-8 -*-
"""
DEIS Cognos Scraper
===================
Descarga los reportes de Atenciones de Urgencia del portal Cognos del DEIS
(Ministerio de Salud de Chile): un archivo .xlsx por establecimiento, con la
serie de años completa. Sirve para cualquier Servicio de Salud del país.

Uso rápido
----------
    python scraper.py                                  menú interactivo
    python scraper.py --servicio "Metropolitano Central"
    python scraper.py --servicio "Ñuble" --visible
    python scraper.py --diagnostico                    inspeccionar la página

Por qué es más largo de lo que parece
-------------------------------------
El visor de Cognos tiene varios comportamientos que hacen fallar la
automatización EN SILENCIO, dejando archivos con el nombre correcto y contenido
equivocado. Los que este script maneja:

  * Los prompts se reinician después de cada informe (vuelven a un solo año, sin
    grupos de edad, con cientos de establecimientos preseleccionados).
  * Las listas hijas (semanas, establecimientos) no se refrescan al cambiar el
    padre: hay que pasar por un ciclo de solicitud.
  * No existe un "Volver" en el visor; el único enlace con ese texto es de la
    cabecera del portal y apunta a un dominio muerto.
  * Hay 10 enlaces "Seleccionar todo" para 5 listas (cinco son de los grupos de
    edad); clickear el equivocado borra selecciones ya hechas.
  * Los prompts se re-renderizan y los id de los <select> cambian.
  * El servidor devuelve errores transitorios (DPR-ERR-2082 y similares).

Por eso, además de aplicar los filtros, el script VERIFICA: comprueba el estado
antes de pedir cada informe y valida el contenido de cada archivo descargado.
Ver README.md para el detalle.

Requisitos
----------
    pip install -r requirements.txt
    playwright install chromium
"""
import argparse
import json
import logging
import os
import re
import sys
import zipfile
import unicodedata
import time
from datetime import datetime
from pathlib import Path

from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout
from rich.console import Console
from rich.panel import Panel
from rich.progress import (Progress, SpinnerColumn, TextColumn, BarColumn,
                           TaskProgressColumn, TimeElapsedColumn)
from rich.prompt import Confirm, Prompt
from rich.table import Table

# ============================================================================
# CONFIGURACIÓN
# ============================================================================
URL_REPORTE = (
    "http://cognos.deis.cl/ibmcognos/cgi-bin/cognos.cgi?b_action=cognosViewer"
    "&ui.action=run&ui.object=/content/folder[@name=%27PUB%27]/folder[@name=%27REPORTES%27]"
    "/folder[@name=%27Atenciones%20de%20Urgencia%27]/report[@name=%27Atenciones%20Urgencia"
    "%20-%20Vista%20por%20semanas%20-%20Servicios%27]&ui.name=Atenciones%20Urgencia%20-%20"
    "Vista%20por%20semanas%20-%20Servicios&run.outputFormat=&run.prompt=true"
)
# Los años NO se fijan acá: se leen del propio reporte al abrirlo, así el script
# sigue sirviendo cuando el DEIS agregue 2026, 2027, etc.

SERVICIOS_DISPONIBLES = [
    "Aconcagua", "Aisén", "Antofagasta", "Araucanía Norte", "Araucanía Sur", "Arauco",
    "Arica", "Atacama", "Bíobío", "Chiloé", "Concepción", "Coquimbo", "Del Maule",
    "Del Reloncaví", "Iquique", "Libertador B. O'Higgins", "Magallanes",
    "Metropolitano Central", "Metropolitano Norte", "Metropolitano Occidente",
    "Metropolitano Oriente", "Metropolitano Sur", "Metropolitano Suroriente", "Ñuble",
    "Osorno", "Talcahuano", "Valdivia", "Valparaíso San Antonio", "Viña Del Mar Quillota",
]
TIPOS_ESTABLECIMIENTO = ["Hospital", "SAPU", "SAR", "SUR", "CEAR", "PAME"]

# Primeras opciones reales de cada lista en Cognos (el orden NO es alfabético).
PISTAS_SERVICIO = ("Talcahuano", "Concepción", "Arauco", "Metropolitano Central",
                   "Aconcagua", "Iquique", "Ñuble", "Chiloé", "Antofagasta")

DIR_DESCARGAS = Path("./descargas")
MAX_REINTENTOS = 3
TIMEOUT_REPORTE = 300_000       # el reporte con 11 años tarda bastante más
TIMEOUT_DESCARGA = 180_000
ESPERA_CARGA_PAGINA = 15

console = Console()
_LOGGER = None

# Códigos de error que devuelve el propio Cognos. Suelen ser transitorios
# (carga del servidor, sesión caída); el administrador del DEIS es el único
# que ve el detalle detrás del SecureErrorID.
CODIGOS_ERROR_COGNOS = [r"DPR-ERR-\d+", r"RSV-[A-Z]{3}-\d+", r"CM-REQ-\d+",
                        r"QE-DEF-\d+", r"CNC-ASV-\d+", r"SecureErrorID"]
PAUSA_DEFECTO = 2.0            # segundos entre establecimientos
ESPERAS_TRAS_ERROR = [30, 90, 180]


class ErrorCognos(RuntimeError):
    """Error devuelto por el servidor de Cognos, no por el script."""


def setup_logging():
    """Un solo handler, aunque se llame varias veces."""
    global _LOGGER
    if _LOGGER is not None:
        return _LOGGER
    DIR_DESCARGAS.mkdir(parents=True, exist_ok=True)
    fmt = logging.Formatter("[%(asctime)s] %(levelname)s - %(message)s", datefmt="%H:%M:%S")
    fh = logging.FileHandler(DIR_DESCARGAS / "log.txt", encoding="utf-8")
    fh.setFormatter(fmt)
    fh.setLevel(logging.DEBUG)
    logger = logging.getLogger("cognos")
    logger.setLevel(logging.DEBUG)
    logger.handlers = [fh]
    _LOGGER = logger
    return logger


def norm(s):
    """minúsculas, sin tildes, sin espacios repetidos: para comparar textos."""
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s.replace("\xa0", " ").lower().strip())


def sanitizar_nombre(nombre):
    nombre = str(nombre).strip()
    nombre = re.sub(r'[<>:"/\\|?*]', "-", nombre)
    nombre = re.sub(r"\s+", " ", nombre)
    return nombre[:110].strip(" .")


def limpiar_pantalla():
    os.system("cls" if os.name == "nt" else "clear")


# ============================================================================
# SCRAPER
# ============================================================================
JS_SELECT_ALL = """
(id) => {
  const sel = document.getElementById(id);
  if (!sel) return {ok:false, motivo:'select no encontrado'};
  let n = 0;
  for (const o of sel.options) { if (o.value !== '') { o.selected = true; n++; } }
  sel.dispatchEvent(new Event('input',  {bubbles:true}));
  sel.dispatchEvent(new Event('change', {bubbles:true}));
  if (typeof sel.onchange === 'function') { try { sel.onchange(); } catch(e) {} }
  return {ok:true, seleccionadas:n, total:sel.options.length};
}
"""

JS_LINK_CERCANO = """
([id, patron, token]) => {
  const sel = document.getElementById(id);
  if (!sel) return null;
  const re = new RegExp(patron, 'i');
  let cont = sel;
  for (let i = 0; i < 8 && cont; i++) {
    cont = cont.parentElement;
    if (!cont) break;
    const a = Array.from(cont.querySelectorAll('a,span[role=link],button'))
      .find(x => re.test((x.textContent || '').trim()));
    if (a) {
      // Marcar ESTE enlace: muchos no tienen id y hay uno igual por cada lista.
      a.setAttribute('data-cognos-auto', token);
      return {token: token, id: a.id || '', texto: (a.textContent||'').trim()};
    }
  }
  return null;
}
"""

# El visor de Cognos va embebido en el portal del DEIS, cuya cabecera tiene
# enlaces con los MISMOS textos ("Volver", "Descargar como Excel") que apuntan
# a www.deis.cl. Tomar el primero que calce saca al navegador del sitio.
JS_CONTROL = r"""
([patron, token]) => {
  const re = new RegExp(patron, 'i');
  const cands = Array.from(document.querySelectorAll(
      'a,button,input[type=button],input[type=submit]'));
  let mejor = null, mejorPuntaje = -1;
  for (const e of cands) {
    const txt = ((e.textContent || e.value || '') + '').replace(/\s+/g, ' ').trim();
    if (!re.test(txt)) continue;
    const href = e.getAttribute('href') || '';
    let externo = false;
    if (/^https?:/i.test(href)) {
      try { externo = new URL(href, location.href).host !== location.host; }
      catch (err) { externo = true; }
    }
    if (externo) continue;                       // enlace del portal, no del visor
    let p = 0;
    if (re.source.indexOf('^') === 0 || txt.length < 40) p += 2;
    if (!href || /^javascript:|^#/i.test(href)) p += 3;   // control real de Cognos
    if (e.id) p += 1;
    if (e.offsetParent || e.getClientRects().length) p += 2;   // visible ahora
    if (p > mejorPuntaje) { mejorPuntaje = p; mejor = e; }
  }
  if (!mejor) return null;
  document.querySelectorAll('[data-cognos-ctl="' + token + '"]')
          .forEach(x => x.removeAttribute('data-cognos-ctl'));
  mejor.setAttribute('data-cognos-ctl', token);
  return {tag: mejor.tagName, id: mejor.id || '',
          href: mejor.getAttribute('href') || '',
          texto: ((mejor.textContent || mejor.value || '') + '').trim().slice(0, 40),
          puntaje: mejorPuntaje};
}
"""


JS_INVENTARIO = """
() => {
  const selects = Array.from(document.querySelectorAll('select')).map(s => ({
    id: s.id, name: s.name, multiple: s.multiple, size: s.size,
    n_opciones: s.options.length,
    primeras: Array.from(s.options).slice(0, 8).map(o => o.text.trim()),
    seleccionadas: Array.from(s.selectedOptions).slice(0, 5).map(o => o.text.trim())
  }));
  const enlaces = Array.from(document.querySelectorAll('a,button,input[type=button],input[type=submit]'))
    .map(e => ({tag: e.tagName, id: e.id || '', href: e.getAttribute('href') || '',
                visible: !!(e.offsetParent || e.getClientRects().length),
                texto: (e.textContent || e.value || '').trim().slice(0, 60)}))
    .filter(e => e.texto);
  const checks = Array.from(document.querySelectorAll("input[type=checkbox],input[role=checkbox]"))
    .map(c => ({id: c.id || '', checked: c.checked,
                aria: c.getAttribute('aria-checked')}));
  return {url: location.href, titulo: document.title,
          selects, enlaces: enlaces.slice(0, 80), checkboxes: checks};
}
"""


class CognosScraper(object):
    def __init__(self, config, logger=None):
        self.config = config
        self.headless = not config.get("visible", False)
        self.log = logger or setup_logging()
        self.playwright = None
        self.browser = None
        self.context = None
        self.page = None
        self._select_ids = {}
        self._anios_disponibles = []
        self._descargas = []
        self.total_descargados = 0
        self.total_errores = 0
        self.total_saltados = 0

    # ---------------------------------------------------------------- ciclo
    def iniciar(self):
        self.log.info("Iniciando Chromium (headless=%s)" % self.headless)
        self.playwright = sync_playwright().start()
        self.browser = self.playwright.chromium.launch(
            headless=self.headless, args=["--disable-blink-features=AutomationControlled"])
        self.context = self.browser.new_context(accept_downloads=True,
                                                viewport={"width": 1400, "height": 950})
        # Capturar descargas venga de donde venga (página o popup).
        self.context.on("page", self._enganchar_pagina)
        self.page = self.context.new_page()
        self._enganchar_pagina(self.page)

    def _enganchar_pagina(self, pagina):
        try:
            pagina.on("download", lambda d: self._descargas.append(d))
        except Exception:
            pass

    def cerrar(self):
        try:
            if self.browser:
                self.browser.close()
        finally:
            if self.playwright:
                self.playwright.stop()
        self.log.info("Navegador cerrado")

    def navegar_al_reporte(self):
        self.log.debug("Navegando al reporte")
        self.page.goto(URL_REPORTE, wait_until="domcontentloaded", timeout=90_000)
        self.page.wait_for_timeout(ESPERA_CARGA_PAGINA * 1000)

    # ------------------------------------------------------------ selectores
    JS_INSPECCION = """
() => Array.from(document.querySelectorAll('select')).map(s => ({
    id: s.id || '', n: s.options.length,
    primeras: Array.from(s.options).slice(0, 8).map(o => (o.text||'').trim())
}))
"""

    def anios_del_reporte(self):
        """Años que ofrece el reporte HOY. Se leen de la página, no se asumen."""
        sid = self._select_ids.get("anio")
        if not sid:
            self.descubrir_selectores()
            sid = self._select_ids.get("anio")
        if not sid:
            return []
        textos = self.page.evaluate(
            "(id) => { const e = document.getElementById(id); return e ? "
            "Array.from(e.options).map(o => (o.text||'').trim()) : []; }", sid)
        return sorted({int(t) for t in textos if re.match(r"^(19|20)\d{2}$", t)})

    def descubrir_selectores(self):
        """Clasifica los <select> por su contenido. Una sola llamada al
        navegador: con 644 opciones, ir elemento por elemento cuesta segundos."""
        info = self.page.evaluate(self.JS_INSPECCION)
        ids = {}
        for s_ in info:
            sid, textos = s_["id"], s_["primeras"]
            if not sid or not textos:
                continue
            digitos = [t for t in textos if t.isdigit()]
            if digitos and all(2010 <= int(t) <= 2035 for t in digitos) and \
                    len(digitos) >= max(1, len(textos) - 1):
                ids["anio"] = sid
            elif any("semana" in t.lower() for t in textos):
                ids["semana"] = sid
            elif any(t in TIPOS_ESTABLECIMIENTO for t in textos):
                ids["tipo_est"] = sid
            elif any(t in PISTAS_SERVICIO for t in textos):
                ids["servicio"] = sid
        # El de establecimiento es el que más opciones tiene entre los que sobran.
        asignados = set(ids.values())
        libres = [(s_["n"], s_["id"]) for s_ in info
                  if s_["id"] and s_["id"] not in asignados and s_["n"]]
        if libres:
            ids["establecimiento"] = max(libres)[1]
        self._select_ids = ids
        self.log.debug("Selectores: %s" % ids)
        return ids

    def _exigir(self, nombre):
        if nombre not in self._select_ids:
            raise RuntimeError(
                "No se encontró el selector '%s'. Corre con --diagnostico para ver "
                "qué hay realmente en la página." % nombre)
        return self._select_ids[nombre]

    def _sel(self, nombre):
        return self.page.locator("#%s" % self._exigir(nombre))

    @staticmethod
    def _id_link(sel_id, accion):
        """Cognos nombra los enlaces de cada lista de forma predecible:
        select  PRMT_SV_<sufijo>  ->  link  PRMT_SV_LINK_SELECT_<sufijo>
        Es la vía exacta; hay 10 enlaces 'Seleccionar todo' en la página
        (5 listas + 5 grupos de edad) y buscarlos por texto agarra el que no es."""
        pref = "PRMT_SV_"
        if sel_id.startswith(pref):
            return "PRMT_SV_LINK_%s_%s" % (accion, sel_id[len(pref):])
        return None

    def _click_link_lista(self, sel_id, accion):
        """Clickea 'Seleccionar todo'/'Deseleccionar todo' de ESA lista."""
        lid = self._id_link(sel_id, accion)
        if lid:
            loc = self.page.locator("#%s" % lid)
            if loc.count():
                loc.first.click()
                return "id"
        # Respaldo: subir por el DOM, pero solo mientras el contenedor
        # tenga exactamente UN select (si no, el enlace puede ser de otra lista).
        patron = "seleccionar\\s+todo" if accion == "SELECT" else "deseleccionar\\s+todo"
        token = "auto_%s_%s" % (accion, sel_id[-8:])
        info = self.page.evaluate(JS_LINK_CERCANO, [sel_id, patron, token])
        if info:
            self.page.locator('[data-cognos-auto="%s"]' % token).first.click()
            return "dom"
        return None

    def _control(self, patron, token, obligatorio=True):
        """Localiza un control DEL VISOR (descarta los enlaces del portal DEIS)."""
        info = self.page.evaluate(JS_CONTROL, [patron, token])
        if not info:
            if obligatorio:
                raise RuntimeError(
                    "No se encontró el control '%s' dentro del visor de Cognos. "
                    "Corre --diagnostico." % token)
            return None
        self.log.debug("Control '%s' -> %s" % (token, info))
        return self.page.locator('[data-cognos-ctl="%s"]' % token).first

    def _boton_nueva_solicitud(self):
        # El botón real de Cognos tiene id que empieza con "reprompt".
        loc = self.page.locator("button[id^='reprompt']")
        if loc.count():
            return loc.first
        return self._control("^\\s*nueva\\s+solicitud\\s*$", "nueva_solicitud")

    def _link_excel(self):
        return self._control("descargar\\s+como\\s+excel", "excel")

    def _click_en_cognos(self, locator, descripcion):
        """Clickea y comprueba que no nos sacó del visor."""
        antes = self.page.url
        locator.click()
        self.page.wait_for_timeout(1500)
        ahora = self.page.url
        if "cognos" not in ahora.lower() and ahora != antes:
            self.log.error("'%s' navegó fuera de Cognos: %s" % (descripcion, ahora))
            try:
                self.page.go_back(wait_until="domcontentloaded", timeout=30_000)
                self.page.wait_for_timeout(2000)
            except Exception:
                pass
            raise RuntimeError(
                "El clic en '%s' salió del visor hacia %s (es un enlace del portal "
                "del DEIS, no un control del reporte)." % (descripcion, ahora))

    def _esperar_informe(self, timeout_ms=None):
        self._link_excel().wait_for(state="visible", timeout=timeout_ms or TIMEOUT_REPORTE)
        self.page.wait_for_timeout(1500)

    def error_de_cognos(self):
        """Devuelve el texto del error si la página muestra uno de Cognos."""
        try:
            txt = self.page.evaluate(
                "() => (document.body ? document.body.innerText : '').slice(0, 6000)")
        except Exception:
            return None
        for pat in CODIGOS_ERROR_COGNOS:
            m = re.search(pat, txt or "", re.I)
            if m:
                ini = max(0, m.start() - 120)
                return re.sub(r"\s+", " ", txt[ini:m.end() + 160]).strip()
        return None

    def _hay_prompts(self):
        """¿Siguen los filtros en pantalla?"""
        try:
            ids = self.descubrir_selectores()
            return bool(ids.get("anio") and ids.get("establecimiento"))
        except Exception:
            return False

    def ejecutar_solicitud(self, timeout_ms=None):
        """Pulsa 'Nueva solicitud' y espera a que el informe termine.

        En este reporte del DEIS los filtros y el panel de Herramientas conviven
        en la misma página: NO hay un 'Volver' al que regresar (el único enlace
        con ese texto es el del portal, http://www.deis.cl/?p=41, que ya no
        resuelve). Así que basta con volver a pedir.
        """
        timeout_ms = timeout_ms or TIMEOUT_REPORTE
        url_antes = self.page.url
        self._boton_nueva_solicitud().click()
        # networkidle cubre los dos casos: recarga completa (postback) y AJAX.
        # Mientras Cognos genera el informe la petición sigue abierta, así que
        # esto espera lo que haga falta en vez de dormir un rato fijo.
        try:
            self.page.wait_for_load_state("networkidle", timeout=timeout_ms)
        except PlaywrightTimeout:
            self.log.warning("Red aún activa tras %d s; se continúa" % (timeout_ms / 1000))
        except Exception as exc:
            self.log.debug("wait_for_load_state: %s" % exc)
        self.page.wait_for_timeout(2500)
        ahora = self.page.url
        if "cognos" not in ahora.lower() and ahora != url_antes:
            raise RuntimeError("Tras 'Nueva solicitud' el navegador terminó en %s" % ahora)
        err = self.error_de_cognos()
        if err:
            raise ErrorCognos(err)
        self.log.info("Solicitud ejecutada (url=%s)" % ahora[:70])
        if not self._hay_prompts():
            # Algunos despliegues muestran solo el informe: recargar el reporte.
            self.log.warning("Los filtros ya no están en pantalla; recargando el reporte")
            self.navegar_al_reporte()
            self.descubrir_selectores()
            return False
        return True

    def _dejar_un_establecimiento(self):
        """Deja exactamente UNO seleccionado y devuelve su nombre."""
        sid = self._select_ids.get("establecimiento")
        if not sid:
            return None
        nombre = self.page.evaluate("""(id) => {
            const e = document.getElementById(id);
            if (!e) return null;
            let elegido = null;
            for (const o of e.options) {
                if (!elegido && o.value !== '') { o.selected = true; elegido = o.text.trim(); }
                else { o.selected = false; }
            }
            e.dispatchEvent(new Event('change', {bubbles:true}));
            return elegido;
        }""", sid)
        self.page.wait_for_timeout(500)
        return nombre

    def ciclo_de_refresco(self):
        """Cognos no repuebla las listas hijas al cambiar el padre: lo hace en el
        siguiente ciclo de solicitud. Corre un informe mínimo (un establecimiento,
        sin grupos de edad) y vuelve, para que la lista de semanas refleje los
        años elegidos."""
        elegido = self._dejar_un_establecimiento()
        console.print("   [dim]Consulta intermedia (1 establecimiento: %s) para que "
                      "Cognos repueble las listas…[/dim]" % (elegido or "?"))
        self.log.info("Ciclo de refresco con 1 establecimiento: %s" % elegido)
        if not self.ejecutar_solicitud():
            raise RuntimeError(
                "Tras la consulta intermedia los filtros desaparecieron de la página; "
                "hubo que recargar el reporte y se perdieron las selecciones.")

    def _n_opciones(self, nombre):
        sid = self._select_ids.get(nombre)
        if not sid:
            return 0
        return self.page.evaluate(
            "(id) => { const e = document.getElementById(id);"
            " return e ? e.options.length : 0; }", sid)

    def esperar_cascada(self, nombre, n_antes, timeout_ms=30_000, exigir_cambio=False,
                        espera_minima_ms=5_000):
        """Espera a que una lista hija termine de repoblarse.

        exigir_cambio=True: además de estable, el tamaño tiene que ser distinto
        al de antes (es lo que se espera al pasar de 1 año a 11). Si no cambia,
        agota el tiempo y devuelve lo que haya, para que el que llama decida.
        """
        inicio = time.time()
        limite = inicio + timeout_ms / 1000.0
        minimo = inicio + espera_minima_ms / 1000.0
        anterior, estable = None, 0
        sid = self._select_ids.get(nombre)
        while time.time() < limite:
            self.page.wait_for_timeout(700)
            if not sid:
                self.descubrir_selectores()
                sid = self._select_ids.get(nombre)
            n = self.page.evaluate(
                "(id) => { const e = document.getElementById(id);"
                " return e ? e.options.length : 0; }", sid) if sid else 0
            if n == 0:                       # el prompt se re-renderizó
                self.descubrir_selectores()
                sid = self._select_ids.get(nombre)
                anterior, estable = None, 0
                continue
            # No aceptar "estable" antes de la espera mínima: el refresco de
            # Cognos tarda, y devolver el valor viejo pasa desapercibido.
            if time.time() < minimo:
                anterior = n
                continue
            if n == anterior and (not exigir_cambio or n != n_antes):
                estable += 1
                if estable >= 2:
                    self.log.debug("Cascada '%s': %s -> %s" % (nombre, n_antes, n))
                    return n
            else:
                estable = 0
            anterior = n
        self.log.warning("Cascada '%s' no cambió (%s)" % (nombre, anterior))
        return anterior or 0


    def seleccionar_todo(self, nombre):
        """Marca todas las opciones de una lista."""
        sel_id = self._exigir(nombre)
        try:
            via = self._click_link_lista(sel_id, "SELECT")
            if via:
                self.page.wait_for_timeout(1200)
                n = self._sel(nombre).evaluate("el => el.selectedOptions.length")
                if n > 1:
                    self.log.info("'%s': %d opciones (enlace por %s)" % (nombre, n, via))
                    return n
        except Exception as exc:
            self.log.warning("Enlace 'Seleccionar todo' de '%s' falló: %s" % (nombre, exc))
        r = self.page.evaluate(JS_SELECT_ALL, sel_id)
        self.page.wait_for_timeout(1200)
        self.log.info("'%s': %s (por JS)" % (nombre, r))
        return r.get("seleccionadas", 0) if isinstance(r, dict) else 0

    JS_ELEGIR = r"""
([id, buscado]) => {
  const norm = s => (s||'').normalize('NFD').replace(/[\u0300-\u036f]/g,'')
                       .toLowerCase().replace(/\s+/g,' ').trim();
  const sel = document.getElementById(id);
  if (!sel) return {estado:'sin_select'};
  const b = norm(buscado);
  const op = Array.from(sel.options).find(o => norm(o.text) === b)
          || Array.from(sel.options).find(o => norm(o.text).includes(b));
  if (!op) return {estado:'sin_opcion',
                   opciones: Array.from(sel.options).slice(0,10).map(o=>o.text.trim())};
  for (const o of sel.options) o.selected = false;
  op.selected = true;
  sel.dispatchEvent(new Event('input',  {bubbles:true}));
  sel.dispatchEvent(new Event('change', {bubbles:true}));
  if (typeof sel.onchange === 'function') { try { sel.onchange(); } catch(e) {} }
  return {estado:'ok', elegido: op.text.trim()};
}
"""

    def seleccionar_por_texto(self, nombre, texto):
        """Deja seleccionada solo la opción que coincide con 'texto'."""
        r = self.page.evaluate(self.JS_ELEGIR, [self._exigir(nombre), texto])
        if r["estado"] == "sin_select":
            # Cognos re-renderiza los prompts: los ids cambian bajo los pies.
            self.log.warning("El select '%s' ya no existe; redescubriendo" % nombre)
            self.descubrir_selectores()
            r = self.page.evaluate(self.JS_ELEGIR, [self._exigir(nombre), texto])
        if r["estado"] != "ok":
            raise RuntimeError(
                "No se pudo elegir '%s' en la lista '%s' (%s). Primeras opciones: %s"
                % (texto, nombre, r["estado"], r.get("opciones", [])[:8]))
        self.page.wait_for_timeout(1200)
        return r["elegido"]

    def marcar_grupos_edad(self):
        """Marca los 5 grupos. Vienen DESMARCADOS al cargar la página."""
        estado = {"total": 0, "marcados": 0}
        for _ in range(2):
            estado = self.page.evaluate("""() => {
                const cbs = Array.from(document.querySelectorAll(
                    "input[type=checkbox],input[role=checkbox]"));
                let clicks = 0;
                for (const c of cbs) { if (!c.checked) { c.click(); clicks++; } }
                return {total: cbs.length, clicks: clicks,
                        marcados: cbs.filter(c => c.checked).length};
            }""")
            self.page.wait_for_timeout(900)
            if estado["total"] and estado["marcados"] == estado["total"]:
                break
        self.log.info("Grupos de edad: %s" % estado)
        if estado["total"] and estado["marcados"] != estado["total"]:
            raise RuntimeError("Solo %d de %d grupos de edad quedaron marcados"
                               % (estado["marcados"], estado["total"]))
        return estado

    # ------------------------------------------------------------- filtros
    def diagnostico(self, ruta=None):
        inv = self.page.evaluate(JS_INVENTARIO)
        ruta = ruta or (DIR_DESCARGAS / "diagnostico_cognos.json")
        ruta.parent.mkdir(parents=True, exist_ok=True)
        with open(str(ruta), "w", encoding="utf-8") as fh:
            json.dump(inv, fh, ensure_ascii=False, indent=1)
        console.print("\n[bold]Selectores encontrados en la página:[/bold]")
        for s in inv["selects"]:
            console.print("  id=[cyan]%s[/cyan] múltiple=%s opciones=%d\n     %s"
                          % (s["id"], s["multiple"], s["n_opciones"],
                             " | ".join(s["primeras"][:5])))
        console.print("\n[bold]Enlaces/botones:[/bold] %s"
                      % ", ".join(sorted({e["texto"] for e in inv["enlaces"]})[:25]))
        console.print("\n[green]Detalle completo -> %s[/green]" % ruta)
        return inv

    def diagnostico_flujo(self):
        """Fotografía la página antes y después de 'Nueva solicitud'. Es lo que
        hace falta para entender el ciclo de este reporte sin adivinar."""
        pasos = []

        def foto(etiqueta):
            inv = self.page.evaluate(JS_INVENTARIO)
            inv["etiqueta"] = etiqueta
            inv["iframes"] = self.page.evaluate(
                "() => Array.from(document.querySelectorAll('iframe'))"
                ".map(f => f.getAttribute('src') || '(sin src)')")
            pasos.append(inv)
            visibles = [e["texto"] for e in inv["enlaces"] if e["visible"]][:6]
            console.print("   [cyan]%-20s[/cyan] selects=%d iframes=%d  visibles: %s"
                          % (etiqueta, len(inv["selects"]), len(inv["iframes"]),
                             ", ".join(visibles)))
            for s_ in inv["selects"]:
                console.print("      %-34s %d opciones" % (s_["id"][:34], s_["n_opciones"]))

        foto("al cargar")
        self.descubrir_selectores()
        if self.config["servicio"] != "TODOS":
            try:
                self.seleccionar_por_texto("servicio", self.config["servicio"])
            except Exception as exc:
                console.print("   [yellow]servicio: %s[/yellow]" % exc)
        elegido = self._dejar_un_establecimiento()
        console.print("   [dim]establecimiento de prueba: %s[/dim]" % elegido)
        foto("filtros minimos")

        console.print("   [dim]pulsando 'Nueva solicitud'…[/dim]")
        self.ejecutar_solicitud()
        foto("tras la solicitud")
        self.page.wait_for_timeout(8000)
        foto("8 s despues")

        ruta = DIR_DESCARGAS / "diagnostico_flujo.json"
        ruta.parent.mkdir(parents=True, exist_ok=True)
        with open(str(ruta), "w", encoding="utf-8") as fh:
            json.dump(pasos, fh, ensure_ascii=False, indent=1)
        console.print("\n[green]Detalle -> %s[/green]" % ruta)
        return pasos

    def aplicar_filtros(self, anios):
        """Aplica años, servicio y tipo; fuerza UN ciclo de solicitud para que
        Cognos repueble las listas hijas; y recién ahí elige semanas y edades.

        El orden importa. Las listas de semanas y de establecimientos son
        prompts en cascada que Cognos NO refresca al cambiar el padre: espera
        al siguiente ciclo de solicitud. Por eso conviene dejar puestos todos
        los filtros "padre" primero y pagar UN solo ciclo, en vez de uno por
        cascada. Ese informe intermedio se pide con un solo establecimiento y
        sin grupos de edad, para que sea lo más liviano posible.
        """
        self.descubrir_selectores()
        if not getattr(self, "_anios_disponibles", None):
            self._anios_disponibles = self.anios_del_reporte()
        n_sem_antes = self._n_opciones("semana")
        n_est_antes = self._n_opciones("establecimiento")

        # 1. Años
        if len(anios) == 1:
            self.log.info("Año: %s" % self.seleccionar_por_texto("anio", str(anios[0])))
        elif sorted(anios) == sorted(self._anios_disponibles or anios):
            self.seleccionar_todo("anio")
        else:
            self.page.evaluate("""([id, lista]) => {
                const sel = document.getElementById(id);
                const q = new Set(lista.map(String));
                for (const o of sel.options) o.selected = q.has(o.text.trim());
                sel.dispatchEvent(new Event('change', {bubbles:true}));
            }""", [self._exigir("anio"), [str(a) for a in anios]])
            self.page.wait_for_timeout(1500)

        # 2. Servicio de Salud
        if self.config["servicio"] != "TODOS":
            real = self.seleccionar_por_texto("servicio", self.config["servicio"])
            console.print("   [dim]Servicio: %s[/dim]" % real)
        else:
            self.seleccionar_todo("servicio")

        # 3. Tipo de establecimiento
        if self.config["tipo_est"] != "TODOS":
            self.seleccionar_por_texto("tipo_est", self.config["tipo_est"])
        else:
            self.seleccionar_todo("tipo_est")

        # 4. Un solo ciclo de solicitud que refresca AMBAS cascadas
        n_sem = self.esperar_cascada("semana", n_sem_antes, timeout_ms=12_000,
                                     exigir_cambio=True, espera_minima_ms=3_000)
        n_est = self._n_opciones("establecimiento")
        hay_que_refrescar = (len(anios) > 1 and n_sem <= n_sem_antes) or \
                            (self.config["servicio"] != "TODOS" and n_est >= n_est_antes)
        if hay_que_refrescar:
            self.ciclo_de_refresco()
            n_sem = self.esperar_cascada("semana", n_sem_antes, timeout_ms=40_000,
                                         exigir_cambio=(len(anios) > 1))
            n_est = self.esperar_cascada(
                "establecimiento", n_est_antes, timeout_ms=30_000,
                exigir_cambio=(self.config["servicio"] != "TODOS"))
        console.print("   [dim]Semanas: %d (antes %d) · Establecimientos: %d (antes %d)[/dim]"
                      % (n_sem, n_sem_antes, n_est, n_est_antes))
        if len(anios) > 1 and n_sem <= n_sem_antes:
            raise RuntimeError(
                "La lista de semanas quedó en %d tras elegir %d años: trae las de un "
                "solo año y el informe saldría incompleto. Manda descargas/log.txt."
                % (n_sem, len(anios)))
        if self.config["servicio"] != "TODOS" and n_est >= n_est_antes:
            console.print("   [yellow]Aviso: la lista de establecimientos no se redujo "
                          "(%d). Puede que traiga todo el país.[/yellow]" % n_est)

        # 5. La consulta intermedia también REINICIA las selecciones (los años
        #    vuelven a 2024, el servicio queda vacío, las edades desmarcadas).
        #    Las cascadas ya están bien, así que basta con volver a elegir todo.
        self.reaplicar_filtros(anios)

        # 6. Comprobar que nada quedó vacío
        self.verificar_filtros()

        # 8. Lista final
        ests = self.page.evaluate(
            "(id) => { const e = document.getElementById(id); return e ? "
            "Array.from(e.options).map(o => ({nombre:o.text.trim(), value:o.value}))"
            ".filter(o => o.nombre !== '') : []; }", self._exigir("establecimiento"))
        self.log.info("Establecimientos disponibles: %d" % len(ests))
        return ests

    def verificar_filtros(self):
        """Comprueba que cada lista tenga algo seleccionado. Un filtro vacío
        genera un reporte vacío sin que Cognos avise."""
        estado = {}
        for nombre in ("anio", "semana", "servicio", "tipo_est"):
            sid = self._select_ids.get(nombre)
            if not sid:
                continue
            estado[nombre] = self.page.evaluate(
                "(id) => { const e = document.getElementById(id);"
                " return e ? [e.selectedOptions.length, e.options.length] : [0,0]; }", sid)
        self.log.info("Estado de los filtros: %s" % estado)
        vacios = [k for k, (sel, _) in estado.items() if sel == 0]
        detalle = "  ".join("%s=%d/%d" % (k, v[0], v[1]) for k, v in estado.items())
        console.print("   [dim]Filtros: %s[/dim]" % detalle)
        if vacios:
            raise RuntimeError(
                "Estos filtros quedaron SIN selección: %s. El reporte saldría vacío. "
                "Corre con --diagnostico para inspeccionar la página." % ", ".join(vacios))

    # ------------------------------------------------------------ descarga
    def _esperar_descarga(self, timeout_ms):
        limite = time.time() + timeout_ms / 1000.0
        while time.time() < limite:
            if self._descargas:
                return self._descargas.pop(0)
            self.page.wait_for_timeout(300)
        return None

    @staticmethod
    def es_xlsx(ruta):
        """Un .xlsx es un ZIP con xl/workbook.xml dentro. Una página de error
        guardada con esa extensión no lo es."""
        try:
            with zipfile.ZipFile(str(ruta)) as z:
                return any(n.startswith("xl/") for n in z.namelist())
        except Exception:
            return False

    @staticmethod
    def analizar_excel(ruta):
        """Lee un export de Cognos y devuelve qué contiene realmente:
        establecimientos del pie, años presentes y bloques etarios."""
        try:
            import openpyxl
        except ImportError:
            return None
        try:
            wb = openpyxl.load_workbook(str(ruta), data_only=True)
            ws = wb.active
            g = lambda r, c: ws.cell(row=r, column=c).value
            nfil, ncol = ws.max_row, ws.max_column

            anios = set()
            for r in range(1, min(nfil, 400)):
                for c in range(2, min(ncol, 800) + 1):
                    v, abajo = g(r, c), g(r + 1, c)
                    if v is None or abajo is None:
                        continue
                    if re.match(r"^(19|20)\d{2}$", str(v).strip()) and \
                            str(abajo).strip().lower() == "total":
                        anios.add(int(str(v).strip()))

            bloques = sum(
                1 for r in range(1, nfil + 1)
                if isinstance(g(r, 1), str) and
                norm(g(r, 1)).startswith("atenciones de urgencia por causa"))

            ests = []
            for r in range(1, nfil + 1):
                v = g(r, 1)
                if isinstance(v, str) and norm(v).startswith("filtros aplicados"):
                    for rr in range(r + 1, nfil + 1):
                        t = g(rr, 1)
                        if t is None or not str(t).strip():
                            continue
                        t = str(t).strip()
                        if norm(t) in ("establecimiento", "servicio de salud",
                                       "ano estadistico", "semana estadistica",
                                       "tipo de establecimiento"):
                            continue
                        ests.append(t)
                    break
            wb.close()
        except Exception:
            return None
        return {"anios": sorted(anios), "bloques_etarios": bloques,
                "establecimientos": ests, "filas": nfil, "columnas": ncol}

    def verificar_archivo(self, ruta, esperado, anios):
        """Comprueba que el archivo sea de ESE establecimiento, con TODOS los años
        y los 6 bloques etarios. Sin esto, un informe corrido con los filtros
        perdidos se guarda con el nombre correcto y pasa desapercibido."""
        if not self.es_xlsx(ruta):
            raise RuntimeError("Archivo inválido: no es un Excel "
                               "(¿página de error de Cognos guardada como .xlsx?)")
        info = self.analizar_excel(ruta)
        if info is None:
            return None
        self.log.info("%s -> %s" % (os.path.basename(str(ruta)),
                                    {k: v for k, v in info.items()
                                     if k != "establecimientos"}))
        problemas = []
        ests = info["establecimientos"]
        if len(ests) > 1:
            problemas.append("trae %d establecimientos sumados (%s…)"
                             % (len(ests), ", ".join(ests[:3])))
        elif len(ests) == 1 and norm(ests[0]) != norm(esperado):
            problemas.append("es de '%s' y no de '%s'" % (ests[0], esperado))
        if info["bloques_etarios"] < 6:
            problemas.append("solo %d de 6 bloques etarios (faltan grupos de edad)"
                             % info["bloques_etarios"])
        faltan = sorted(set(anios) - set(info["anios"]))
        if faltan:
            problemas.append("le faltan los años %s (trae %s)"
                             % (faltan, info["anios"] or "ninguno"))
        if problemas:
            raise RuntimeError("Archivo inválido: " + "; ".join(problemas))
        return True

    def _estado(self):
        return self.page.evaluate("""(ids) => {
            const out = {};
            for (const k in ids) {
                const e = ids[k] && document.getElementById(ids[k]);
                out[k] = e ? [e.selectedOptions.length, e.options.length] : [0, 0];
            }
            const cbs = Array.from(document.querySelectorAll(
                "input[type=checkbox],input[role=checkbox]"));
            out.edades = [cbs.filter(c => c.checked).length, cbs.length];
            return out;
        }""", {k: self._select_ids.get(k) for k in
               ("anio", "semana", "servicio", "tipo_est", "establecimiento")})

    def reaplicar_filtros(self, anios):
        """Vuelve a poner años, servicio, tipo, edades y semanas.

        Cognos reinicia los prompts a su estado por defecto (2024, sin grupos de
        edad, con medio país seleccionado) después de generar un informe. Si no
        se re-aplican, el siguiente archivo sale con TODO sumado y con el nombre
        del establecimiento correcto: indetectable a simple vista.
        """
        try:
            self.page.wait_for_load_state("networkidle", timeout=30_000)
        except Exception:
            pass
        self.page.wait_for_timeout(1200)
        self.descubrir_selectores()

        if len(anios) == 1:
            self.seleccionar_por_texto("anio", str(anios[0]))
        else:
            self.seleccionar_todo("anio")
        if self.config["servicio"] != "TODOS":
            self.seleccionar_por_texto("servicio", self.config["servicio"])
        else:
            self.seleccionar_todo("servicio")
        if self.config["tipo_est"] != "TODOS":
            self.seleccionar_por_texto("tipo_est", self.config["tipo_est"])
        else:
            self.seleccionar_todo("tipo_est")
        self.marcar_grupos_edad()
        self.descubrir_selectores()

        # Si la lista de semanas volvió a la de un solo año, hace falta el ciclo.
        n_sem = self._n_opciones("semana")
        if len(anios) > 1 and n_sem <= 60:
            self.log.warning("Las semanas volvieron a %d: nuevo ciclo de refresco" % n_sem)
            self.ciclo_de_refresco()
            self.esperar_cascada("semana", n_sem, timeout_ms=40_000, exigir_cambio=True)
            self.marcar_grupos_edad()
            self.descubrir_selectores()
        self.seleccionar_todo("semana")

    def descargar_establecimiento(self, est, ruta_destino, anios, verificar=True):
        self.reaplicar_filtros(anios)

        # Dejar solo este establecimiento seleccionado. Al cargar, Cognos trae
        # varios marcados; si quedara más de uno la data saldría sumada.
        sel_id = self._exigir("establecimiento")
        try:
            self._click_link_lista(sel_id, "DESELECT")
            self.page.wait_for_timeout(400)
        except Exception:
            pass
        self.page.evaluate("""([id, val]) => {
            const sel = document.getElementById(id);
            for (const o of sel.options) o.selected = (o.value === val);
            sel.dispatchEvent(new Event('change', {bubbles:true}));
        }""", [sel_id, est["value"]])
        self.page.wait_for_timeout(600)
        # Comprobación completa justo antes de pedir el informe.
        est_actual = self._estado()
        self.log.info("Estado antes de solicitar: %s" % est_actual)
        fallos = []
        if est_actual["establecimiento"][0] != 1:
            fallos.append("establecimientos seleccionados=%d (debe ser 1)"
                          % est_actual["establecimiento"][0])
        if est_actual["anio"][0] != len(anios):
            fallos.append("años=%d de %d" % (est_actual["anio"][0], len(anios)))
        if est_actual["semana"][0] == 0:
            fallos.append("sin semanas")
        if est_actual["edades"][0] != est_actual["edades"][1]:
            fallos.append("grupos de edad %d/%d" % tuple(est_actual["edades"]))
        if self.config["servicio"] != "TODOS" and est_actual["servicio"][0] != 1:
            fallos.append("servicios=%d (debe ser 1)" % est_actual["servicio"][0])
        if fallos:
            raise RuntimeError("Los filtros no quedaron bien: " + "; ".join(fallos))
        console.print("      [dim]años=%d semanas=%d edades=%d/%d est=1[/dim]"
                      % (est_actual["anio"][0], est_actual["semana"][0],
                         est_actual["edades"][0], est_actual["edades"][1]))

        self.ejecutar_solicitud()
        self._esperar_informe()

        ruta_destino.parent.mkdir(parents=True, exist_ok=True)
        self._descargas = []
        self._link_excel().click()
        descarga = self._esperar_descarga(TIMEOUT_DESCARGA)
        if descarga is None:
            raise RuntimeError("No llegó ninguna descarga tras pulsar 'Descargar como Excel'")
        descarga.save_as(str(ruta_destino))

        for p in list(self.context.pages):
            if p is not self.page:
                try:
                    p.close()
                except Exception:
                    pass

        if not (ruta_destino.exists() and ruta_destino.stat().st_size > 0):
            return False
        if not verificar:                 # solo para las pruebas con el mock
            self.descubrir_selectores()
            return True
        try:
            comprobado = self.verificar_archivo(ruta_destino, est["nombre"], anios)
        except RuntimeError:
            try:
                ruta_destino.unlink()     # no dejar el archivo malo en disco
            except Exception:
                pass
            raise
        if comprobado is None:
            self.log.debug("No se pudo comprobar el contenido de %s" % ruta_destino.name)

        # Los filtros siguen en pantalla: solo hay que releer los ids.
        self.descubrir_selectores()
        return True


# ============================================================================
# UI INTERACTIVA
# ============================================================================
def mostrar_banner():
    console.print(Panel.fit(
        "[bold cyan]DEIS COGNOS · Atenciones de Urgencia[/bold cyan]\n"
        "[dim]Descarga automática desde cognos.deis.cl[/dim]",
        border_style="cyan"))


def solicitar_servicio():
    console.print("\n[bold yellow]PASO 1 · SERVICIO DE SALUD[/bold yellow]")
    console.print("  [green][0][/green] TODOS LOS SERVICIOS (Chile completo)")
    for i, s in enumerate(SERVICIOS_DISPONIBLES, 1):
        console.print("  [cyan][%2d][/cyan] %s" % (i, s))
    while True:
        opc = Prompt.ask("\n➤ Número o nombre")
        if opc == "0" or opc.lower() in ("todos", "all"):
            return "TODOS"
        if opc.isdigit() and 1 <= int(opc) <= len(SERVICIOS_DISPONIBLES):
            return SERVICIOS_DISPONIBLES[int(opc) - 1]
        coincid = [s for s in SERVICIOS_DISPONIBLES if opc.lower() in s.lower()]
        if len(coincid) == 1:
            return coincid[0]
        console.print("[red]Entrada inválida o ambigua: %s[/red]" % (coincid or ""))


def solicitar_anios():
    console.print("\n[bold yellow]PASO 2 · AÑOS ESTADÍSTICOS[/bold yellow]")
    console.print("  [cyan][1][/cyan] Todos los que ofrezca el reporte "
                  "[green]en un solo archivo por establecimiento[/green]")
    console.print("  [cyan][2][/cyan] Un año específico")
    console.print("  [cyan][3][/cyan] Rango de años")
    opc = Prompt.ask("➤ Opción", choices=["1", "2", "3"], default="1")
    if opc == "1":
        return "todos"
    if opc == "2":
        while True:
            a = Prompt.ask("➤ Año")
            if re.match(r"^(19|20)\d{2}$", a):
                return [int(a)]
            console.print("[red]Escribe un año de 4 dígitos.[/red]")
    while True:
        r = Prompt.ask("➤ Rango (ej: 2020-2024)")
        m = re.match(r"^\s*(\d{4})\s*-\s*(\d{4})\s*$", r)
        if m:
            d, h = int(m.group(1)), int(m.group(2))
            if d <= h:
                return list(range(d, h + 1))
        console.print("[red]Formato incorrecto o fuera de rango.[/red]")


def solicitar_tipo_est():
    console.print("\n[bold yellow]PASO 3 · TIPO DE ESTABLECIMIENTO[/bold yellow]")
    console.print("  [cyan][1][/cyan] Todos (%s)" % ", ".join(TIPOS_ESTABLECIMIENTO))
    console.print("  [cyan][2][/cyan] Uno específico")
    if Prompt.ask("➤ Opción", choices=["1", "2"], default="1") == "1":
        return "TODOS"
    for i, t in enumerate(TIPOS_ESTABLECIMIENTO, 1):
        console.print("  [cyan][%d][/cyan] %s" % (i, t))
    while True:
        idx = Prompt.ask("➤ Número")
        if idx.isdigit() and 1 <= int(idx) <= len(TIPOS_ESTABLECIMIENTO):
            return TIPOS_ESTABLECIMIENTO[int(idx) - 1]
        console.print("[red]Número inválido.[/red]")


def main_interactivo():
    limpiar_pantalla()
    mostrar_banner()
    config = {"servicio": solicitar_servicio()}
    config["anios"] = solicitar_anios()
    config["tipo_est"] = solicitar_tipo_est()
    config["establecimiento"] = "TODOS"
    config["establecimientos"] = None
    config["visible"] = Confirm.ask("\n➤ ¿Mostrar el navegador mientras trabaja?",
                                    default=False)
    config["pausa"] = PAUSA_DEFECTO

    t = Table(title="RESUMEN", border_style="green")
    t.add_column("Parámetro", style="bold")
    t.add_column("Valor", style="cyan")
    t.add_row("Servicio", config["servicio"])
    t.add_row("Años", "todos los disponibles" if config["anios"] == "todos"
              else ("%s-%s" % (config["anios"][0], config["anios"][-1])
                    if len(config["anios"]) > 1 else str(config["anios"][0])))
    t.add_row("Semanas", "Todas")
    t.add_row("Grupos de edad", "Los 5")
    t.add_row("Tipo establecimiento", config["tipo_est"])
    t.add_row("Establecimientos", "Todos, uno por archivo")
    console.print()
    console.print(t)

    if Confirm.ask("\n[bold green]¿Iniciar descarga?[/bold green]", default=True):
        ejecutar(config)
    else:
        console.print("[yellow]Cancelado.[/yellow]")


# ============================================================================
# EJECUCIÓN
# ============================================================================
def ejecutar(config):
    logger = setup_logging()
    scraper = CognosScraper(config, logger)
    anios, destino = [], DIR_DESCARGAS

    console.print("\n[bold cyan]Iniciando…[/bold cyan]")
    try:
        with console.status("[bold blue]Abriendo Chromium…[/bold blue]"):
            scraper.iniciar()
        with console.status("[bold blue]Cargando el reporte…[/bold blue]"):
            scraper.navegar_al_reporte()

        # Los años se toman del reporte, no de una lista fija en el código.
        scraper.descubrir_selectores()
        disponibles = scraper.anios_del_reporte()
        scraper._anios_disponibles = disponibles
        if not disponibles:
            console.print("[red]No se pudo leer la lista de años. Corre --diagnostico.[/red]")
            return
        pedidos = config["anios"]
        if pedidos in (None, "todos"):
            anios = list(disponibles)
        else:
            anios = [a for a in pedidos if a in disponibles]
            fuera = [a for a in pedidos if a not in disponibles]
            if fuera:
                console.print("[yellow]El reporte no ofrece: %s. Disponibles: %d-%d[/yellow]"
                              % (fuera, disponibles[0], disponibles[-1]))
        if not anios:
            console.print("[red]Ninguno de los años pedidos existe en el reporte.[/red]")
            return
        etiqueta = ("%d-%d" % (anios[0], anios[-1])) if len(anios) > 1 else str(anios[0])
        destino = DIR_DESCARGAS / sanitizar_nombre(config["servicio"]) / etiqueta
        console.print("   [dim]Años del reporte: %d-%d · se usarán %d[/dim]"
                      % (disponibles[0], disponibles[-1], len(anios)))

        if config.get("diagnostico") or config.get("diagnostico_flujo"):
            scraper.descubrir_selectores()
            scraper.diagnostico()
            if config.get("diagnostico_flujo"):
                console.print("\n[bold]Flujo de la solicitud[/bold]")
                scraper.diagnostico_flujo()
            return

        with console.status("[bold yellow]Aplicando filtros (esto tarda)…[/bold yellow]"):
            ests = scraper.aplicar_filtros(anios)

        if config.get("establecimientos"):
            objetivos = [norm(e) for e in config["establecimientos"]]
            ests = [e for e in ests
                    if any(o in norm(e["nombre"]) or norm(e["nombre"]) in o
                           for o in objetivos)]
            console.print("   [dim]Filtro --establecimientos: %d seleccionados[/dim]"
                          % len(ests))
        if not ests:
            console.print("[red]No se obtuvo ningún establecimiento. "
                          "Corre con --diagnostico.[/red]")
            return
        console.print("[green]%d establecimiento(s) por descargar -> %s[/green]\n"
                      % (len(ests), destino))

        with Progress(SpinnerColumn(), TextColumn("[progress.description]{task.description}"),
                      BarColumn(complete_style="green"), TaskProgressColumn(),
                      TimeElapsedColumn(), console=console) as prog:
            tarea = prog.add_task("Descargando…", total=len(ests))
            for est in ests:
                nombre = ("Atenciones Urgencia - Vista por semanas - Servicios - %s.xlsx"
                          % sanitizar_nombre(est["nombre"]))
                ruta = destino / nombre
                if ruta.exists() and ruta.stat().st_size > 0:
                    # No basta con que exista: hay que comprobar que sea válido.
                    # Una corrida anterior pudo dejar archivos con los filtros
                    # perdidos, con el nombre correcto y datos de todo el país.
                    try:
                        scraper.verificar_archivo(ruta, est["nombre"], anios)
                        scraper.total_saltados += 1
                        prog.console.print("[dim]· ya está: %s[/dim]" % est["nombre"])
                        prog.advance(tarea)
                        continue
                    except RuntimeError as exc:
                        prog.console.print("[yellow]· rehaciendo %s -> %s[/yellow]"
                                           % (est["nombre"][:38], str(exc)[:70]))
                        try:
                            ruta.unlink()
                        except Exception:
                            pass

                prog.update(tarea, description="[cyan]%s[/cyan]" % est["nombre"][:45])
                exito = False
                for intento in range(1, MAX_REINTENTOS + 1):
                    try:
                        exito = scraper.descargar_establecimiento(est, ruta, anios)
                        if exito:
                            break
                    except Exception as exc:
                        logger.error("Intento %d en '%s': %s" % (intento, est["nombre"], exc))
                        if isinstance(exc, ErrorCognos):
                            espera = ESPERAS_TRAS_ERROR[min(intento - 1,
                                                            len(ESPERAS_TRAS_ERROR) - 1)]
                            prog.console.print(
                                "[yellow]  el servidor de Cognos devolvió un error "
                                "(no es del script). Esperando %ds antes de reintentar…"
                                "[/yellow]\n  [dim]%s[/dim]" % (espera, str(exc)[:110]))
                            time.sleep(espera)
                        else:
                            prog.console.print("[yellow]  intento %d/%d falló: %s[/yellow]"
                                               % (intento, MAX_REINTENTOS, str(exc)[:90]))
                        if intento < MAX_REINTENTOS:
                            # Recuperación completa: recargar el reporte y rehacer
                            # todos los filtros desde cero.
                            try:
                                scraper.navegar_al_reporte()
                                scraper.aplicar_filtros(anios)
                            except Exception as exc2:
                                logger.error("No se pudo recuperar: %s" % exc2)
                if exito:
                    scraper.total_descargados += 1
                    prog.console.print("[green]OK  %s[/green]" % nombre)
                else:
                    scraper.total_errores += 1
                    prog.console.print("[red]FALLO  %s[/red]" % est["nombre"])
                prog.advance(tarea)
                time.sleep(config.get("pausa", PAUSA_DEFECTO))

    except KeyboardInterrupt:
        console.print("\n[red]Cancelado por el usuario.[/red]")
    except Exception as exc:
        console.print("\n[red]ERROR: %s[/red]" % exc)
        logger.exception("Error fatal")
        console.print("[dim]Detalle en %s[/dim]" % (DIR_DESCARGAS / "log.txt"))
    finally:
        scraper.cerrar()
    if config.get("diagnostico") or config.get("diagnostico_flujo"):
        return
    console.print()
    console.print(Panel(
        "Descargados: [green]%d[/green]\nSaltados:    [yellow]%d[/yellow]\n"
        "Errores:     [red]%d[/red]\nCarpeta:     %s"
        % (scraper.total_descargados, scraper.total_saltados,
           scraper.total_errores, destino),
        title="Resumen", border_style="cyan", expand=False))


def parsear_anios(txt):
    if not txt or txt.lower() in ("todos", "all"):
        return "todos"          # se resuelve al abrir el reporte
    m = re.match(r"^\s*(\d{4})\s*-\s*(\d{4})\s*$", txt)
    if m:
        return list(range(int(m.group(1)), int(m.group(2)) + 1))
    return [int(x) for x in re.split(r"[,\s]+", txt.strip()) if x.isdigit()]


def main():
    ap = argparse.ArgumentParser(description="Descarga automática de reportes COGNOS DEIS")
    ap.add_argument("--servicio", default=None,
                    help="Servicio de Salud (ej: \"Metropolitano Central\") o TODOS. "
                         "Sin este argumento se abre el menú interactivo.")
    ap.add_argument("--anios", default="todos",
                    help="'todos' (los que ofrezca el reporte), '2024', "
                         "'2020-2024' o '2021,2023'")
    ap.add_argument("--tipo", default="TODOS", help="Hospital, SAPU, SAR, SUR, CEAR, PAME")
    ap.add_argument("--visible", action="store_true", help="Mostrar el navegador")
    ap.add_argument("--diagnostico", action="store_true",
                    help="Volcar los selectores reales de la página y salir")
    ap.add_argument("--diagnostico-flujo", action="store_true",
                    help="Fotografiar la página antes y después de 'Nueva solicitud'")
    ap.add_argument("--establecimientos", nargs="+", default=None,
                    help="Bajar solo estos establecimientos (coincidencia parcial)")
    ap.add_argument("--pausa", type=float, default=PAUSA_DEFECTO,
                    help="Segundos entre establecimientos (por defecto %.0f). "
                         "Súbelo si el servidor devuelve errores." % PAUSA_DEFECTO)
    ap.add_argument("--interactivo", action="store_true", help="Menú paso a paso")
    args = ap.parse_args()

    # Sin argumentos, o sin --servicio: menú interactivo (es lo más amable
    # para quien lo usa por primera vez).
    if args.interactivo or (not args.servicio and not args.diagnostico
                            and not args.diagnostico_flujo):
        try:
            main_interactivo()
        except KeyboardInterrupt:
            console.print("\n[yellow]Salida.[/yellow]")
        return

    ejecutar({"servicio": args.servicio or "TODOS",
              "anios": parsear_anios(args.anios),
              "tipo_est": args.tipo, "establecimiento": "TODOS",
              "visible": args.visible or args.diagnostico or args.diagnostico_flujo,
              "diagnostico": args.diagnostico,
              "diagnostico_flujo": args.diagnostico_flujo,
              "establecimientos": args.establecimientos,
              "pausa": args.pausa})


if __name__ == "__main__":
    main()
