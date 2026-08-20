# -*- coding: utf-8 -*-
"""Prueba la lógica del scraper contra una réplica local de la página de Cognos."""
import os, sys, json, pathlib
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault("PLAYWRIGHT_BROWSERS_PATH", "/opt/pw-browsers")
import scraper as C

MOCK = "file://" + str((pathlib.Path(__file__).parent / "mock_cognos.html").resolve())
C.URL_REPORTE = MOCK
C.ESPERA_CARGA_PAGINA = 1
C.DIR_DESCARGAS = pathlib.Path(__file__).parent / "_test_out"

ANIOS = list(range(2015, 2026))
fallos = []
def chk(cond, msg):
    print(("  OK   " if cond else "  FALLA ") + msg)
    if not cond: fallos.append(msg)

cfg = {"servicio": "Metropolitano Central", "anios": list(ANIOS),
       "tipo_est": "TODOS", "establecimiento": "TODOS", "visible": False}
s = C.CognosScraper(cfg)
s.iniciar()
try:
    s.navegar_al_reporte()

    print("\n[1] Descubrimiento de selectores")
    ids = s.descubrir_selectores()
    chk(ids.get("anio") == "PRMT_SV_N16FFCEC0x18BF3DBC_NS_", "año -> %s" % ids.get("anio"))
    chk(ids.get("semana") == "PRMT_SV_N16FFCEC0x18BF4158_NS_", "semana -> %s" % ids.get("semana"))
    chk(ids.get("servicio") == "PRMT_SV_N16FFCEC0x3557F5A0_NS_", "servicio -> %s" % ids.get("servicio"))
    chk(True, "(los prompts se re-renderizan luego: los ids cambian a _R2_)")
    chk(ids.get("tipo_est") == "PRMT_SV_N16FFCEC0x3557DBB8_NS_", "tipo -> %s" % ids.get("tipo_est"))
    chk(ids.get("establecimiento") == "PRMT_SV_N16FFCEC0x3557DED0_NS_",
        "establecimiento -> %s (el de 644 opciones)" % ids.get("establecimiento"))

    print("\n[2] Diagnóstico")
    inv = s.diagnostico(C.DIR_DESCARGAS / "diag.json")
    chk(len(inv["selects"]) == 5, "5 selects inventariados")
    n_links = len([e for e in inv["enlaces"] if e["texto"] == "Seleccionar todo"])
    chk(n_links == 10, "10 enlaces 'Seleccionar todo' como en el real (%d)" % n_links)
    chk(len(inv["checkboxes"]) == 5, "5 checkboxes inventariados")

    print("\n[3] Filtros completos (todos los años)")
    s._anios_disponibles = s.anios_del_reporte()
    chk(s._anios_disponibles == ANIOS, "años leídos del reporte: %s" % (s._anios_disponibles[:3],))
    ests = s.aplicar_filtros(ANIOS)
    def sel_de(nombre, expr):
        return s.page.evaluate("(id) => { const e = document.getElementById(id); return e ? (%s) : null; }" % expr,
                               s._select_ids[nombre])
    est_sel = sel_de("anio", "Array.from(e.selectedOptions).map(o=>o.text)")
    chk(len(est_sel) == 11, "11 años seleccionados a la vez (%d)" % len(est_sel))
    nsem = sel_de("semana", "e.selectedOptions.length")
    chk(nsem == 11*52, "todas las semanas seleccionadas (%d)" % nsem)
    nopt = sel_de("semana", "e.options.length")
    chk(nopt == 11*52, "esperó la cascada del servidor: %d opciones de semana" % nopt)
    ncb = s.page.evaluate("() => Array.from(document.querySelectorAll('#prompts input[type=checkbox]')).filter(c=>c.checked).length")
    chk(ncb == 5, "5 grupos de edad marcados (%d)" % ncb)
    serv = sel_de("servicio", "Array.from(e.selectedOptions).map(o=>o.text)")
    chk(serv == ["Metropolitano Central"], "solo Metropolitano Central: %s" % serv)
    chk(len(ests) == 11, "solo los 11 de Met. Central, no los 644 (%d)" % len(ests))
    chk(any("Maipú" in e["nombre"] for e in ests), "SAPU Maipú presente")
    chk(any("Consultorio Nº1" in e["nombre"] for e in ests), "SAPU Consultorio Nº1 presente")

    chk("mock_cognos.html" in s.page.url, "nunca salió del visor (url: %s)" % s.page.url[-30:])

    print("\n[4] Descarga de 3 establecimientos")
    for est in ests[:3]:
        ruta = C.DIR_DESCARGAS / ("%s.xlsx" % C.sanitizar_nombre(est["nombre"]))
        ok = s.descargar_establecimiento(est, ruta, ANIOS, verificar=False)
        payload = json.loads(ruta.read_text(encoding="utf-8")) if ruta.exists() else {}
        chk(ok and ruta.exists() and ruta.stat().st_size > 0, "descargó %s" % est["nombre"][:40])
        chk(payload.get("est") == [est["nombre"]],
            "  el reporte pidió SOLO ese establecimiento: %s" % payload.get("est"))
        chk(len(payload.get("anios", [])) == 11, "  con los 11 años (%d)" % len(payload.get("anios", [])))
        chk(payload.get("semanas", 0) == 572, "  con las 572 semanas (%d)" % payload.get("semanas", 0))
        chk(payload.get("edades") == 5, "  con los 5 grupos de edad")

    print("\n[5] Normaliza la selección aunque queden varios marcados")
    s.page.evaluate("(id) => { const el=document.getElementById(id);"
                    "for(const o of el.options) o.selected=true; }", s._select_ids["establecimiento"])
    ruta = C.DIR_DESCARGAS / "solo_uno.xlsx"
    s.descargar_establecimiento(ests[0], ruta, ANIOS, verificar=False)
    p5 = json.loads(ruta.read_text(encoding="utf-8"))
    chk(p5.get("est") == [ests[0]["nombre"]],
        "fuerza un solo establecimiento aunque estuvieran todos: %s" % p5.get("est"))

    print("\n[6] La verificación detecta un filtro vacío")
    s.page.evaluate("(id) => { const el=document.getElementById(id);"
                    "for(const o of el.options) o.selected=false; }", s._select_ids["semana"])
    try:
        s.verificar_filtros()
        chk(False, "debió detectar semanas en 0")
    except RuntimeError as e:
        chk("semana" in str(e), "detecta el filtro vacío: %s" % str(e)[:60])

    print("\n[7] Validación de los archivos descargados")

    def excel_cognos(ruta, establecimientos, anios, bloques):
        """Arma un .xlsx con la estructura de un export real de Cognos."""
        import openpyxl
        wb = openpyxl.Workbook(); ws = wb.active
        ws.cell(row=1, column=1,
                value="Atenciones de Urgencia - Vista por Semanas Estadísticas")
        fila = 3
        for b in range(bloques):
            ws.cell(row=fila, column=1,
                    value="Atenciones de Urgencia por Causa y Semanas Estadísticas"
                          + ("" if b == 0 else " - grupo %d" % b))
            col = 2
            for a in anios:
                ws.cell(row=fila + 1, column=col, value=str(a))
                ws.cell(row=fila + 2, column=col, value="Total")
                for w in range(1, 4):
                    ws.cell(row=fila + 2, column=col + w, value=w)
                col += 4
            for i, causa in enumerate(["TOTAL DEMANDA",
                                       "SECCIÓN 1. TOTAL ATENCIONES DE URGENCIA",
                                       "IRA Alta (J00-J06)"]):
                ws.cell(row=fila + 3 + i, column=1, value=causa)
                ws.cell(row=fila + 3 + i, column=2, value=100 + i)
            fila += 10
        ws.cell(row=fila, column=1, value="Filtros aplicados")
        ws.cell(row=fila + 1, column=1, value="Establecimiento")
        for j, e in enumerate(establecimientos):
            ws.cell(row=fila + 2 + j, column=1, value=e)
        wb.save(str(ruta)); wb.close()

    tmp = C.DIR_DESCARGAS; tmp.mkdir(parents=True, exist_ok=True)
    A11 = list(range(2015, 2026))

    bueno = tmp / "bueno.xlsx"
    excel_cognos(bueno, ["SAPU Maipú"], A11, 6)
    try:
        chk(s.verificar_archivo(bueno, "SAPU Maipú", A11) is True,
            "acepta un export completo")
    except RuntimeError as e:
        chk(False, "rechazó uno bueno: %s" % e)

    for nom, ests_f, anios_f, bl, frag in [
            ("sumado",  ["SAPU Maipú", "Hospital X", "SAPU Y"], A11, 6, "sumados"),
            ("otro",    ["Hospital X"], A11, 6, "es de"),
            ("un_anio", ["SAPU Maipú"], [2024], 6, "faltan los años"),
            ("1bloque", ["SAPU Maipú"], A11, 1, "bloques etarios")]:
        f = tmp / ("malo_%s.xlsx" % nom)
        excel_cognos(f, ests_f, anios_f, bl)
        try:
            s.verificar_archivo(f, "SAPU Maipú", A11)
            chk(False, "debió rechazar '%s'" % nom)
        except RuntimeError as e:
            chk(frag in str(e), "rechaza '%s': %s" % (nom, str(e)[:52]))

    nox = tmp / "no_es_excel.xlsx"
    nox.write_text("<html><body>DPR-ERR-2082 ...</body></html>", encoding="utf-8")
    try:
        s.verificar_archivo(nox, "SAPU Maipú", A11)
        chk(False, "debió rechazar la página de error")
    except RuntimeError as e:
        chk("no es un Excel" in str(e), "rechaza una página de error guardada como .xlsx")

    print("\n[8] Detección de errores del servidor")
    for txt, esperado in [("DPR-ERR-2082 Se ha producido un error", True),
                          ("RSV-BBP-0022 the requested session does not exist", True),
                          ("Atenciones de Urgencia - Vista por Semanas", False)]:
        s.page.evaluate("(t) => { document.body.innerHTML = '<p>' + t + '</p>'; }", txt)
        got = s.error_de_cognos()
        chk(bool(got) == esperado,
            "%-46s -> %s" % (txt[:46], "detectado" if got else "sin error"))
finally:
    s.cerrar()

print("\n[9] Funciones puras")

chk(C.parsear_anios("todos") == "todos", "parsear_anios('todos') -> marcador dinámico")
chk(C.parsear_anios("2020-2024") == [2020,2021,2022,2023,2024], "parsear_anios('2020-2024')")
chk(C.parsear_anios("2021,2023") == [2021,2023], "parsear_anios('2021,2023')")
chk(C.parsear_anios("2024") == [2024], "parsear_anios('2024')")
chk(C.sanitizar_nombre("SAPU Dr. Iván Insunza") == "SAPU Dr. Iván Insunza", "sanitizar conserva acentos")
chk("/" not in C.sanitizar_nombre("a/b:c*d"), "sanitizar quita caracteres inválidos")

print("\n%s  %d fallos" % ("TODO OK" if not fallos else "HAY FALLOS", len(fallos)))
for f in fallos: print("   -", f)
sys.exit(1 if fallos else 0)
